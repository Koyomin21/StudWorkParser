from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from string import ascii_uppercase
import os
from openpyxl import Workbook



def save_txt(path,content):
    with open(path,'w',encoding='utf-8')as file:
        for cont in content:
            file.write(cont)

def save_table(file_name,columns,content):
    workbook = Workbook()
    sheet = workbook.active

    for i,case in enumerate(ascii_uppercase):
        if len(columns) == i:
            break
        sheet.cell(row=1,column=i+1).value = columns[i]
        sheet.cell(row=2,column=i+1).value = content[0][i]


        

    workbook.save(filename=file_name)


def files_exist(files):
    downloads = os.listdir('files/')
    print(f'downloads {downloads}')
    print(f'Files: {files}')
    exists = False
    for download in downloads:
        for file in files:
            if file in download:
                exists = True
        if exists:
            exists = False
        else:
            return False

    return True
    
PATH_TO_CHROMEDRIVER = 'C:\\chromedriver.exe'
PATH_TO_YOUR_CHROME_ACCOUNT = r'C:\Users\Ануар\AppData\Local\Google\Chrome\User Data\Profile 1'

chrome_data_dir = PATH_TO_YOUR_CHROME_ACCOUNT
URL = 'https://studwork.org/orders'
PATH = PATH_TO_CHROMEDRIVER

download_directory=os.getcwd()+'\\files\\'
print(download_directory)
prefs = {'download.default_directory':download_directory}

options = webdriver.ChromeOptions()
options.add_argument('user-data-dir='+chrome_data_dir)#set a user
options.add_experimental_option('prefs',prefs) #change the dowload directory



driver = webdriver.Chrome(PATH,options=options)


if not os.path.exists('orders/'):
    os.mkdir('orders/')

if not os.path.exists('files/'):
    os.mkdir('files')


parsed_orders = []#container for parsed orders
cats = ['Название','Раздел','Срок сдачи','Цена','Предмет','Антиплагиат','Тип работы','Описание работы']

orders = driver.find_elements_by_class_name('order-item')


for page in range(1,2):
    driver.get(URL+'?page={}'.format(page))
    base_window = driver.current_window_handle#current page with orders
    orders = driver.find_elements_by_class_name('order-item')


    for order in orders:
        link = order.find_element_by_tag_name('a')
        link.click()

        driver.switch_to.window(driver.window_handles[-1])

        #getting basic categories
        cells = driver.find_elements_by_class_name('chess-board-cell')
        card = [driver.title]
        for cell in cells:
            cat = cell.text.split('\n')
            if cat[0] in cats:
                card.append(cat[-1])
        
        parsed_orders.append(card)
        #get order number
        order_number = driver.find_element_by_class_name('chess-board-cell__value')
        print(order_number.text)
        #getting the description
        description = driver.find_element_by_xpath('//*[@id="app"]/div[3]/div/div[1]/div[2]/div/div/div[2]/div[2]/div').text
        parsed_orders[-1].append(description)
        
        #getting files
        file_names = []
        files = driver.find_elements_by_class_name('file-info.file__info')
        for file in files:
            link_text = file.text.split('\n')[0]
            print(link_text)
            try:
                link = driver.find_element_by_link_text(link_text)
                link.click() #download the file
                file_names.append(link_text)
            except:
                pass

        
        
        if not os.path.exists('orders\\{}\\'.format(order_number.text)):
            os.mkdir('orders\\{}\\'.format(order_number.text))

        file_name = 'orders\\{order_number}\\{order_number}.xlsx'.format(order_number=order_number.text)
        save_table(file_name,cats,parsed_orders)
        save_txt('orders\\{order_number}\\links.txt'.format(order_number=order_number.text),file_names)

        driver.close()#close current window
        driver.switch_to_window(base_window)#switch to main window(orders page)
        parsed_orders.clear()




    
print('Success!')
driver.quit()

